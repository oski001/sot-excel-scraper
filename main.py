import re
import sys
import requests
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

URLS = {
    "chests": "https://seaofthieves.fandom.com/wiki/Treasure_Chests",
    "skulls": "https://seaofthieves.fandom.com/wiki/Bounty_Skulls",
    "athena": "https://seaofthieves.fandom.com/wiki/Athena%27s_Fortune_Treasure"
}

RANGE_RE = re.compile(r"(\d[\d,]*)\s*[–-]\s*(\d[\d,]*)")  # hyphen or en-dash
ON_BOARD_LOOT_MULTIPLIER = 0  # Change if needed


def median_from_range(text: str):
    m = RANGE_RE.search(text)
    if not m:
        text = text.replace(",", "").strip()
        if text.isdigit():
            return float(text)
        return None
    lo, hi = (int(x.replace(",", "")) for x in m.groups())
    return (lo + hi) / 2


def scrape_table(table, name_col=0, median_col=1):
    """Scrape name and median from specified columns."""
    rows = []
    for tr in table.find_all("tr")[1:]:
        tds = tr.find_all("td")
        if len(tds) <= max(name_col, median_col):
            continue
        name = tds[name_col].get_text(strip=True)
        median = median_from_range(tds[median_col].get_text(strip=True))
        if median is not None:
            rows.append((name, median))
    return rows


def safe_sheet(name):
    bad = set(r'[]:*?/\\')
    return "".join(c for c in name if c not in bad)[:31] or "Sheet"


def autofit(ws):
    for col in ws.columns:
        max_length = max((len(str(c.value)) if c.value else 0) for c in col) + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length


def table_for_heading(h3):
    nxt = h3.find_next_sibling(
        lambda t: t.name == "table" and "wikitable" in t.get("class", [])
    )
    prv = h3.find_previous_sibling(
        lambda t: t.name == "table" and "wikitable" in t.get("class", [])
    )
    if not nxt:
        return prv
    if not prv:
        return nxt
    return nxt if len(list(h3.next_elements)) < len(list(h3.previous_elements)) else prv


def add_sheet(wb, title, table, name_col=0, median_col=1):
    data = scrape_table(table, name_col=name_col, median_col=median_col)
    if not data:
        return None

    ws = wb.create_sheet(safe_sheet(title))
    ws.append(["Name", "Median Base Gold Reward", "On-Board Loot", "Total Gold"])

    for idx, (name, median) in enumerate(data, start=2):
        ws.append([name, median, ON_BOARD_LOOT_MULTIPLIER])
        ws[f"D{idx}"] = f"=B{idx}*C{idx}"

    total_rows = ws.max_row
    sum_cell = f"D{total_rows + 1}"
    ws[sum_cell] = f"=SUM(D2:D{total_rows})"
    ws[sum_cell].font = Font(bold=True)

    autofit(ws)
    return data


def main(heading_id=None):
    print("⛵ Fetching pages…")
    wb = Workbook()
    wb.remove(wb.active)
    master = wb.create_sheet("All Medians")
    master.append(["Category", "Name", "Median Base Gold Reward", "On-Board Loot", "Total Gold"])

    row_index = 2

    for category, url in URLS.items():
        print(f"Fetching {category} from {url} ...")
        soup = BeautifulSoup(requests.get(url, timeout=30).text, "html.parser")

        if category == "athena":
            athena_tables = [
                table for table in soup.find_all("table", class_="wikitable")
                if any("Base" in th.get_text(strip=True) and "Gold" in th.get_text(strip=True) for th in table.find_all("th"))
            ]

            for i, table in enumerate(athena_tables, start=1):
                sheet_title = f"{category} Treasure {i}"
                data = add_sheet(wb, sheet_title, table)
                if data:
                    for name, median in data:
                        master.append([category, name, median, ON_BOARD_LOOT_MULTIPLIER])
                        master[f"E{row_index}"] = f"=C{row_index}*D{row_index}"
                        row_index += 1

        else:
            # Existing chests and skulls logic unchanged
            for h3 in soup.find_all("h3"):
                span = h3.find("span", class_="mw-headline")
                if not span:
                    continue
                table = table_for_heading(h3)
                if not table:
                    continue
                headers = [th.get_text(strip=True) for th in table.find_all("th")]
                if not any("Base" in h and "Gold" in h for h in headers):
                    continue
                data = add_sheet(wb, f"{category} - {span.get_text(strip=True)}", table)
                if data:
                    for name, median in data:
                        master.append([category, name, median, ON_BOARD_LOOT_MULTIPLIER])
                        master[f"E{row_index}"] = f"=C{row_index}*D{row_index}"
                        row_index += 1

    total_rows = master.max_row
    sum_cell = f"E{total_rows + 1}"
    master[sum_cell] = f"=SUM(E2:E{total_rows})"
    master[sum_cell].font = Font(bold=True)

    autofit(master)

    outfile = "sea_of_thieves_treasure_all.xlsx"
    wb.save(outfile)
    print(f"Finished — saved all data to {Path(outfile).resolve()}")


if __name__ == "__main__":
    heading_arg = sys.argv[1] if len(sys.argv) > 1 else None
    main(heading_arg)
